"""Inventory risk snapshot lead capture for outbound campaigns."""
from __future__ import annotations

import csv
import io
import os
from collections import Counter
from datetime import datetime
from typing import Annotated, Literal, Optional
from urllib.parse import urlparse

from fastapi import APIRouter, Depends, Header, HTTPException, Response, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from pydantic import BaseModel, EmailStr, Field, field_validator
from sqlalchemy import select
from sqlalchemy.orm import Session as DbSession

from app.db.models import InventoryRiskSnapshotLead
from app.db.session import get_db_session

router = APIRouter(prefix="/inventory-risk-snapshot", tags=["inventory-risk-snapshot"])

InventoryIssue = Literal[
    "Stockouts",
    "Overstock / dead stock",
    "Reorder planning",
    "Supplier lead times",
    "Bundles / kits",
    "Not sure",
]

LeadStatus = Literal["New", "Reviewing", "Snapshot sent", "Demo booked", "Not qualified"]


def _check_admin_token(token: Optional[str]) -> None:
    expected = os.getenv("ADMIN_BOOTSTRAP_TOKEN")
    if not expected:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Admin export is disabled (ADMIN_BOOTSTRAP_TOKEN not set).",
        )
    if not token or token != expected:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Invalid admin token.")


def _normalize_store_url(value: str) -> str:
    trimmed = value.strip().lower()
    if not trimmed:
        raise ValueError("Shopify store URL is required.")
    if " " in trimmed or "." not in trimmed:
        raise ValueError("Enter a valid Shopify store URL.")

    parsed = urlparse(trimmed if "://" in trimmed else f"https://{trimmed}")
    host = parsed.netloc or parsed.path
    if not host or "." not in host:
        raise ValueError("Enter a valid Shopify store URL.")
    if host in {"example.com", "test.com", "localhost"} or host.endswith(".test"):
        raise ValueError("Enter your real Shopify store URL.")
    return f"https://{host}{parsed.path if parsed.netloc else ''}".rstrip("/")


class SnapshotLeadRequest(BaseModel):
    first_name: str = Field(min_length=1, max_length=120)
    email: EmailStr
    company_name: str = Field(min_length=1, max_length=255)
    store_url: str = Field(min_length=4, max_length=500)
    approximate_sku_count: str = Field(min_length=1, max_length=64)
    biggest_inventory_issue: InventoryIssue
    source: str = Field(default="inventory_risk_snapshot", max_length=96)
    utm_source: Optional[str] = Field(default=None, max_length=255)
    utm_medium: Optional[str] = Field(default=None, max_length=255)
    utm_campaign: Optional[str] = Field(default=None, max_length=255)
    utm_content: Optional[str] = Field(default=None, max_length=255)
    utm_term: Optional[str] = Field(default=None, max_length=255)

    @field_validator("first_name", "company_name", "approximate_sku_count", "source")
    @classmethod
    def _strip_required(cls, value: str) -> str:
        stripped = value.strip()
        if not stripped:
            raise ValueError("This field is required.")
        return stripped

    @field_validator("store_url")
    @classmethod
    def _valid_store_url(cls, value: str) -> str:
        return _normalize_store_url(value)

    @field_validator("utm_source", "utm_medium", "utm_campaign", "utm_content", "utm_term")
    @classmethod
    def _strip_optional(cls, value: Optional[str]) -> Optional[str]:
        stripped = (value or "").strip()
        return stripped or None


class SnapshotLeadResponse(BaseModel):
    id: int
    created_at: datetime
    first_name: str
    email: str
    company_name: str
    store_url: str
    approximate_sku_count: str
    biggest_inventory_issue: str
    source: str
    status: str


def _to_response(lead: InventoryRiskSnapshotLead) -> SnapshotLeadResponse:
    return SnapshotLeadResponse(
        id=lead.id,
        created_at=lead.created_at,
        first_name=lead.first_name,
        email=lead.email,
        company_name=lead.company_name,
        store_url=lead.store_url,
        approximate_sku_count=lead.approximate_sku_count,
        biggest_inventory_issue=lead.biggest_inventory_issue,
        source=lead.source,
        status=lead.status,
    )


@router.post("/leads", response_model=SnapshotLeadResponse)
def create_snapshot_lead(
    request: SnapshotLeadRequest,
    db: Annotated[DbSession, Depends(get_db_session)],
) -> SnapshotLeadResponse:
    lead = InventoryRiskSnapshotLead(
        first_name=request.first_name,
        email=str(request.email).lower().strip(),
        company_name=request.company_name,
        store_url=request.store_url,
        approximate_sku_count=request.approximate_sku_count,
        biggest_inventory_issue=request.biggest_inventory_issue,
        source=request.source,
        utm_source=request.utm_source,
        utm_medium=request.utm_medium,
        utm_campaign=request.utm_campaign,
        utm_content=request.utm_content,
        utm_term=request.utm_term,
        status="New",
    )
    db.add(lead)
    db.commit()
    db.refresh(lead)
    return _to_response(lead)


@router.get("/leads", response_model=list[SnapshotLeadResponse])
def list_snapshot_leads(
    db: Annotated[DbSession, Depends(get_db_session)],
    x_admin_token: Annotated[Optional[str], Header(alias="X-Admin-Token")] = None,
    status_filter: Optional[LeadStatus] = None,
    limit: int = 200,
) -> list[SnapshotLeadResponse]:
    _check_admin_token(x_admin_token)
    query = select(InventoryRiskSnapshotLead).order_by(InventoryRiskSnapshotLead.created_at.desc())
    if status_filter:
        query = query.where(InventoryRiskSnapshotLead.status == status_filter)
    leads = db.scalars(query.limit(min(max(limit, 1), 1000))).all()
    return [_to_response(lead) for lead in leads]


@router.get("/leads/export.csv")
def export_snapshot_leads_csv(
    db: Annotated[DbSession, Depends(get_db_session)],
    x_admin_token: Annotated[Optional[str], Header(alias="X-Admin-Token")] = None,
) -> Response:
    """Legacy CSV export — kept as a fallback. Prefer /leads/export.xlsx for
    the formatted workbook with a Summary tab."""
    _check_admin_token(x_admin_token)
    leads = db.scalars(
        select(InventoryRiskSnapshotLead).order_by(InventoryRiskSnapshotLead.created_at.desc())
    ).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "id", "created_at", "first_name", "email", "company_name", "store_url",
        "approximate_sku_count", "biggest_inventory_issue", "source",
        "utm_source", "utm_medium", "utm_campaign", "utm_content", "utm_term", "status",
    ])
    for lead in leads:
        writer.writerow([
            lead.id,
            lead.created_at.isoformat() if lead.created_at else "",
            lead.first_name, lead.email, lead.company_name, lead.store_url,
            lead.approximate_sku_count, lead.biggest_inventory_issue, lead.source,
            lead.utm_source or "", lead.utm_medium or "", lead.utm_campaign or "",
            lead.utm_content or "", lead.utm_term or "", lead.status,
        ])
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=inventory-risk-snapshot-leads.csv"},
    )


@router.get("/leads/export.xlsx")
def export_snapshot_leads_xlsx(
    db: Annotated[DbSession, Depends(get_db_session)],
    x_admin_token: Annotated[Optional[str], Header(alias="X-Admin-Token")] = None,
) -> Response:
    """Formatted xlsx export — Summary tab (KPIs + data-bar charts) plus a
    filterable Leads tab. Same auth model as the CSV endpoint."""
    _check_admin_token(x_admin_token)
    leads = db.scalars(
        select(InventoryRiskSnapshotLead).order_by(InventoryRiskSnapshotLead.created_at.desc())
    ).all()

    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    _build_leads_summary_sheet(summary_ws, leads)

    detail_ws = workbook.create_sheet("Leads")
    _build_leads_detail_sheet(detail_ws, leads)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=inventory-risk-snapshot-leads.xlsx"
        },
    )


# ---------------------------------------------------------------------------
# xlsx formatting — palette matches frontend report-export.ts so customer-
# facing and admin exports share visual language.
# ---------------------------------------------------------------------------
_BRAND = "FF0F172A"
_BRAND_INK = "FFFFFFFF"
_BRAND_MUTED = "FFCBD5E1"
_BORDER = "FFDBE3EF"
_SECTION_BG = "FFEEF4FF"
_HEADER_BG = "FFDBEAFE"
_HEADER_INK = "FF334155"
_ALT_ROW = "FFF8FBFF"
_ACCENT_NEUTRAL = "FF2563EB"
_ACCENT_GOOD = "FF0F766E"
_ACCENT_WARNING = "FFB45309"
_ACCENT_DANGER = "FFB91C1C"

_TONE_FILL = {
    "good": "FFDFF7EF",
    "warning": "FFFFF3CF",
    "danger": "FFFEE2E2",
    "neutral": "FFFFFFFF",
}
_TONE_INK = {
    "good": "FF065F46",
    "warning": "FF92400E",
    "danger": "FF991B1B",
    "neutral": "FF0F172A",
}


def _fill(argb: str) -> PatternFill:
    return PatternFill("solid", fgColor=argb)


def _thin_border() -> Border:
    side = Side(style="thin", color=_BORDER)
    return Border(top=side, bottom=side, left=side, right=side)


def _hair_border() -> Border:
    side = Side(style="hair", color=_BORDER)
    return Border(top=side, bottom=side, left=side, right=side)


def _build_leads_summary_sheet(ws, leads) -> None:
    ws.sheet_view.showGridLines = False
    cols = 12
    for col_idx in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 28 if col_idx == 1 else 12

    # Brand bar
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols - 2)
    brand = ws.cell(row=1, column=1, value="SKUBASE EXPORT")
    brand.font = Font(name="Calibri", bold=True, size=11, color=_BRAND_INK)
    brand.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col_idx in range(1, cols + 1):
        ws.cell(row=1, column=col_idx).fill = _fill(_BRAND)
    ws.merge_cells(start_row=1, start_column=cols - 1, end_row=1, end_column=cols)
    gen = ws.cell(
        row=1,
        column=cols - 1,
        value=f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
    )
    gen.font = Font(name="Calibri", size=10, color=_BRAND_MUTED)
    gen.alignment = Alignment(vertical="center", horizontal="right", indent=1)
    ws.row_dimensions[1].height = 22

    # Title
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    title = ws.cell(row=2, column=1, value="Inventory Risk Snapshot Leads")
    title.font = Font(name="Calibri", bold=True, size=26, color=_BRAND_INK)
    title.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for col_idx in range(1, cols + 1):
        ws.cell(row=2, column=col_idx).fill = _fill(_BRAND)
    ws.row_dimensions[2].height = 42

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=cols)
    subtitle = ws.cell(
        row=3,
        column=1,
        value="Outbound campaign leads captured by the inventory-risk-snapshot offer.",
    )
    subtitle.font = Font(name="Calibri", size=11, color=_BRAND_MUTED)
    subtitle.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
    for col_idx in range(1, cols + 1):
        ws.cell(row=3, column=col_idx).fill = _fill(_BRAND)
    ws.row_dimensions[3].height = 32

    # KPI row
    total = len(leads)
    new_count = sum(1 for lead in leads if lead.status == "New")
    booked = sum(1 for lead in leads if lead.status == "Demo booked")
    qualified = sum(
        1 for lead in leads if lead.status in {"Reviewing", "Snapshot sent", "Demo booked"}
    )

    kpis = [
        ("Total leads", str(total), "neutral"),
        ("New (untouched)", str(new_count), "warning" if new_count else "neutral"),
        ("In pipeline", str(qualified), "good" if qualified else "neutral"),
        ("Demos booked", str(booked), "good" if booked else "neutral"),
    ]

    kpi_header_row = 5
    kpi_value_row = 6
    kpi_note_row = 7
    card_width = cols // len(kpis)

    for idx, (label, value, tone) in enumerate(kpis):
        start_col = idx * card_width + 1
        end_col = cols if idx == len(kpis) - 1 else start_col + card_width - 1
        fill = _fill(_TONE_FILL[tone])
        ink = _TONE_INK[tone]

        ws.merge_cells(
            start_row=kpi_header_row, start_column=start_col,
            end_row=kpi_header_row, end_column=end_col,
        )
        lab = ws.cell(row=kpi_header_row, column=start_col, value=label.upper())
        lab.font = Font(name="Calibri", bold=True, size=9, color="FF64748B")
        lab.alignment = Alignment(vertical="center", horizontal="left", indent=1)

        ws.merge_cells(
            start_row=kpi_value_row, start_column=start_col,
            end_row=kpi_value_row, end_column=end_col,
        )
        val = ws.cell(row=kpi_value_row, column=start_col, value=value)
        val.font = Font(name="Calibri", bold=True, size=22, color=ink)
        val.alignment = Alignment(vertical="center", horizontal="left", indent=1)

        ws.merge_cells(
            start_row=kpi_note_row, start_column=start_col,
            end_row=kpi_note_row, end_column=end_col,
        )

        for r in range(kpi_header_row, kpi_note_row + 1):
            for c in range(start_col, end_col + 1):
                ws.cell(row=r, column=c).fill = fill

    ws.row_dimensions[kpi_header_row].height = 20
    ws.row_dimensions[kpi_value_row].height = 32
    ws.row_dimensions[kpi_note_row].height = 18

    # Charts
    row = kpi_note_row + 2

    def _draw_chart(title: str, points) -> None:
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        section = ws.cell(row=row, column=1, value=title)
        section.font = Font(name="Calibri", bold=True, size=14, color=_BRAND)
        section.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        for c in range(1, cols + 1):
            ws.cell(row=row, column=c).fill = _fill(_SECTION_BG)
        ws.row_dimensions[row].height = 28
        row += 1

        if not points:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
            empty = ws.cell(row=row, column=1, value="(no data)")
            empty.font = Font(name="Calibri", italic=True, color="FF64748B")
            empty.alignment = Alignment(vertical="center", horizontal="left", indent=1)
            row += 1
            return

        max_val = max((p[1] for p in points), default=1) or 1
        for label, value, tone in points:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            lab = ws.cell(row=row, column=1, value=label)
            lab.font = Font(name="Calibri", bold=True, size=11, color=_HEADER_INK)
            lab.alignment = Alignment(vertical="center", horizontal="left", indent=1)

            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=10)
            bar = ws.cell(row=row, column=4, value=value)
            bar.number_format = "#,##0"
            bar.font = Font(name="Calibri", size=1, color="00FFFFFF")

            ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=12)
            disp = ws.cell(row=row, column=11, value=str(value))
            disp.font = Font(name="Calibri", bold=True, size=11, color=_TONE_INK[tone])
            disp.alignment = Alignment(vertical="center", horizontal="right", indent=1)

            accent = {
                "danger": _ACCENT_DANGER,
                "warning": _ACCENT_WARNING,
                "good": _ACCENT_GOOD,
                "neutral": _ACCENT_NEUTRAL,
            }[tone]
            ws.conditional_formatting.add(
                f"D{row}:J{row}",
                DataBarRule(
                    start_type="num", start_value=0,
                    end_type="num", end_value=max_val,
                    color=accent[2:],
                    gradient=True, showValue=False,
                ),
            )

            ws.row_dimensions[row].height = 22
            row += 1

        ws.row_dimensions[row].height = 10
        row += 1

    status_counts = Counter(lead.status for lead in leads)
    status_order = ["New", "Reviewing", "Snapshot sent", "Demo booked", "Not qualified"]
    tone_for_status = {
        "New": "warning",
        "Reviewing": "warning",
        "Snapshot sent": "good",
        "Demo booked": "good",
        "Not qualified": "neutral",
    }
    status_points = [
        (s, status_counts.get(s, 0), tone_for_status[s])
        for s in status_order
        if status_counts.get(s, 0) > 0 or s == "New"
    ]
    _draw_chart("Status mix", status_points)

    issue_counts = Counter(
        lead.biggest_inventory_issue for lead in leads if lead.biggest_inventory_issue
    )
    issue_points = sorted(
        [(issue, count, "neutral") for issue, count in issue_counts.items()],
        key=lambda x: x[1],
        reverse=True,
    )
    _draw_chart("Biggest inventory issue", issue_points)

    source_counts = Counter(lead.utm_source or lead.source or "direct" for lead in leads)
    source_points = sorted(
        [(s, count, "neutral") for s, count in source_counts.items()],
        key=lambda x: x[1],
        reverse=True,
    )[:8]
    _draw_chart("Acquisition source", source_points)

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    pointer = ws.cell(row=row, column=1, value='Full lead list on the "Leads" tab.')
    pointer.font = Font(name="Calibri", italic=True, size=11, color="FF64748B")
    pointer.alignment = Alignment(vertical="center", horizontal="left", indent=1)

    ws.freeze_panes = "A4"


def _build_leads_detail_sheet(ws, leads) -> None:
    ws.sheet_view.showGridLines = False

    headers = [
        ("id", "ID", 8),
        ("created_at", "Captured at", 20),
        ("first_name", "First name", 16),
        ("email", "Email", 30),
        ("company_name", "Company", 24),
        ("store_url", "Store URL", 30),
        ("approximate_sku_count", "SKU count", 12),
        ("biggest_inventory_issue", "Biggest issue", 22),
        ("source", "Source", 14),
        ("utm_source", "UTM source", 14),
        ("utm_medium", "UTM medium", 14),
        ("utm_campaign", "UTM campaign", 18),
        ("utm_content", "UTM content", 16),
        ("utm_term", "UTM term", 14),
        ("status", "Status", 16),
    ]

    for idx, (_key, _label, width) in enumerate(headers):
        ws.column_dimensions[get_column_letter(idx + 1)].width = width

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1, value="Lead Records")
    title.font = Font(name="Calibri", bold=True, size=18, color=_BRAND_INK)
    title.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).fill = _fill(_BRAND)
    ws.row_dimensions[1].height = 32

    header_row = 2
    for idx, (_key, label, _w) in enumerate(headers):
        cell = ws.cell(row=header_row, column=idx + 1, value=label)
        cell.font = Font(name="Calibri", bold=True, size=10, color=_HEADER_INK)
        cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        cell.fill = _fill(_HEADER_BG)
        cell.border = _thin_border()
    ws.row_dimensions[header_row].height = 26

    status_tone = {
        "New": "warning",
        "Reviewing": "warning",
        "Snapshot sent": "good",
        "Demo booked": "good",
        "Not qualified": "neutral",
    }

    for r_idx, lead in enumerate(leads):
        r = header_row + 1 + r_idx
        alt = r_idx % 2 == 1
        row_values = {
            "id": lead.id,
            "created_at": lead.created_at.strftime("%Y-%m-%d %H:%M") if lead.created_at else "",
            "first_name": lead.first_name,
            "email": lead.email,
            "company_name": lead.company_name,
            "store_url": lead.store_url,
            "approximate_sku_count": lead.approximate_sku_count,
            "biggest_inventory_issue": lead.biggest_inventory_issue,
            "source": lead.source,
            "utm_source": lead.utm_source or "",
            "utm_medium": lead.utm_medium or "",
            "utm_campaign": lead.utm_campaign or "",
            "utm_content": lead.utm_content or "",
            "utm_term": lead.utm_term or "",
            "status": lead.status,
        }
        for idx, (key, _label, _w) in enumerate(headers):
            cell = ws.cell(row=r, column=idx + 1, value=row_values.get(key))
            tone = "neutral"
            if key == "status":
                tone = status_tone.get(lead.status, "neutral")
            fill_argb = _TONE_FILL[tone] if tone != "neutral" else (_ALT_ROW if alt else "FFFFFFFF")
            ink = _TONE_INK[tone] if tone != "neutral" else "FF0F172A"
            cell.font = Font(name="Calibri", size=10, bold=tone != "neutral", color=ink)
            cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
            cell.fill = _fill(fill_argb)
            cell.border = _hair_border()
        ws.row_dimensions[r].height = 22

    if leads:
        last_col_letter = get_column_letter(len(headers))
        last_row = header_row + len(leads)
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row}"
    ws.freeze_panes = f"A{header_row + 1}"
